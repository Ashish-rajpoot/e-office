from openpyxl import load_workbook


def get_all_users(file_path, sheet_name="Sheet1"):

    workbook = load_workbook(
        file_path,
        keep_vba=True
    )

    sheet = workbook[sheet_name]

    users = []

    for row in range(2, sheet.max_row + 1):

        user_id = sheet[f"C{row}"].value
        flag = sheet[f"H{row}"].value
        name = sheet[f"B{row}"].value

        if user_id:
            users.append(
                {
                    "row": row,
                    "user_id": user_id,
                    "name": name,
                    "flag": flag
                }
            )

    return users